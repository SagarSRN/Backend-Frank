"""
Retail Materials Views
API endpoints for retail estimation
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from decimal import Decimal

from .models import MaterialCategory, Material, RetailEstimate, Component
from .serializers import (
    MaterialCategorySerializer,
    MaterialSerializer,
    RetailEstimateSerializer,
    ComponentSerializer
)


class MaterialCategoryViewSet(viewsets.ModelViewSet):
    """API endpoint for material categories"""
    queryset = MaterialCategory.objects.all()
    serializer_class = MaterialCategorySerializer


class MaterialViewSet(viewsets.ModelViewSet):
    """API endpoint for materials"""
    queryset = Material.objects.all()
    serializer_class = MaterialSerializer
    
    def get_queryset(self):
        queryset = Material.objects.all()
        category = self.request.query_params.get('category', None)
        
        if category:
            queryset = queryset.filter(category_id=category)
        
        return queryset


class RetailEstimateViewSet(viewsets.ModelViewSet):
    """API endpoint for retail estimates"""
    queryset = RetailEstimate.objects.all()
    serializer_class = RetailEstimateSerializer
    
    def get_queryset(self):
        queryset = RetailEstimate.objects.all()
        project_id = self.request.query_params.get('project_id', None)
        
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def by_category(self, request, pk=None):
        """Get estimate components grouped by category"""
        
        estimate = self.get_object()
        components = estimate.components.select_related('material', 'material__category').all()
        
        # Group by category
        categories_dict = {}
        
        for component in components:
            category_name = component.material.category.name
            category_type = component.material.category.category_type
            
            if category_name not in categories_dict:
                categories_dict[category_name] = {
                    'category_name': category_name,
                    'category_type': category_type,
                    'components': [],
                    'total': Decimal('0.00')
                }
            
            component_data = {
                'id': component.id,
                'description': component.description,
                'material_spec': component.material_spec,
                'quantity': float(component.quantity),
                'unit': component.unit,
                'rate': float(component.rate),
                'amount': float(component.amount)
            }
            
            categories_dict[category_name]['components'].append(component_data)
            categories_dict[category_name]['total'] += component.amount
        
        # Convert to list and format totals
        categories_list = []
        for category_data in categories_dict.values():
            category_data['total'] = float(category_data['total'])
            categories_list.append(category_data)
        
        return Response(categories_list)
    
    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """Download estimate as PDF"""
        
        estimate = self.get_object()
        
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from io import BytesIO
            
            # Create PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#4A5568'),
                spaceAfter=30,
            )
            
            elements.append(Paragraph(f"Retail Display Estimate", title_style))
            elements.append(Paragraph(f"Project: {estimate.project.name}", styles['Heading2']))
            elements.append(Spacer(1, 0.3 * inch))
            
            # Summary table
            summary_data = [
                ['Material Cost:', f"₹{estimate.material_cost:,.2f}"],
                ['Labor Cost:', f"₹{estimate.labor_cost:,.2f}"],
                ['Overhead ({:.0f}%):'.format(estimate.overhead_percentage), f"₹{estimate.overhead_amount:,.2f}"],
                ['Profit ({:.0f}%):'.format(estimate.profit_percentage), f"₹{estimate.profit_amount:,.2f}"],
                ['Subtotal:', f"₹{estimate.subtotal:,.2f}"],
                ['GST ({:.0f}%):'.format(estimate.gst_percentage), f"₹{estimate.gst_amount:,.2f}"],
                ['Total:', f"₹{estimate.total:,.2f}"],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2E8F0')),
            ]))
            
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5 * inch))
            
            # Components by category
            elements.append(Paragraph("Bill of Quantities", styles['Heading2']))
            elements.append(Spacer(1, 0.2 * inch))
            
            components = estimate.components.select_related('material', 'material__category').all()
            categories_dict = {}
            
            for component in components:
                category_name = component.material.category.name
                
                if category_name not in categories_dict:
                    categories_dict[category_name] = []
                
                categories_dict[category_name].append(component)
            
            for category_name, comps in categories_dict.items():
                elements.append(Paragraph(category_name, styles['Heading3']))
                
                comp_data = [['Description', 'Qty', 'Unit', 'Rate', 'Amount']]
                
                category_total = Decimal('0.00')
                for comp in comps:
                    comp_data.append([
                        comp.description,
                        f"{comp.quantity:.2f}",
                        comp.unit,
                        f"₹{comp.rate:.2f}",
                        f"₹{comp.amount:,.2f}"
                    ])
                    category_total += comp.amount
                
                comp_data.append(['', '', '', 'Subtotal:', f"₹{category_total:,.2f}"])
                
                comp_table = Table(comp_data, colWidths=[2.5*inch, 0.8*inch, 0.6*inch, 1*inch, 1.1*inch])
                comp_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A5568')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2E8F0')),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ]))
                
                elements.append(comp_table)
                elements.append(Spacer(1, 0.3 * inch))
            
            # Build PDF
            doc.build(elements)
            
            # Return response
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="estimate_{estimate.project.name}.pdf"'
            
            return response
            
        except ImportError:
            return Response(
                {'error': 'PDF generation not available. Install reportlab: pip install reportlab'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {'error': f'PDF generation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def download_excel(self, request, pk=None):
        """Download estimate as Excel"""
        
        estimate = self.get_object()
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Estimate"
            
            # Header
            ws['A1'] = "Retail Display Estimate"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Project: {estimate.project.name}"
            ws['A2'].font = Font(size=12, bold=True)
            
            # Summary
            row = 4
            ws[f'A{row}'] = "Material Cost:"
            ws[f'B{row}'] = float(estimate.material_cost)
            row += 1
            
            ws[f'A{row}'] = "Labor Cost:"
            ws[f'B{row}'] = float(estimate.labor_cost)
            row += 1
            
            ws[f'A{row}'] = f"Overhead ({estimate.overhead_percentage}%):"
            ws[f'B{row}'] = float(estimate.overhead_amount)
            row += 1
            
            ws[f'A{row}'] = f"Profit ({estimate.profit_percentage}%):"
            ws[f'B{row}'] = float(estimate.profit_amount)
            row += 1
            
            ws[f'A{row}'] = "Subtotal:"
            ws[f'B{row}'] = float(estimate.subtotal)
            row += 1
            
            ws[f'A{row}'] = f"GST ({estimate.gst_percentage}%):"
            ws[f'B{row}'] = float(estimate.gst_amount)
            row += 1
            
            ws[f'A{row}'] = "Total:"
            ws[f'B{row}'] = float(estimate.total)
            ws[f'B{row}'].font = Font(bold=True)
            row += 2
            
            # Components
            ws[f'A{row}'] = "Bill of Quantities"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 2
            
            # Headers
            headers = ['Description', 'Quantity', 'Unit', 'Rate', 'Amount']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            row += 1
            
            # Components data
            components = estimate.components.select_related('material').all()
            
            for component in components:
                ws.cell(row=row, column=1, value=component.description)
                ws.cell(row=row, column=2, value=float(component.quantity))
                ws.cell(row=row, column=3, value=component.unit)
                ws.cell(row=row, column=4, value=float(component.rate))
                ws.cell(row=row, column=5, value=float(component.amount))
                row += 1
            
            # Format currency
            for r in range(4, row):
                ws[f'B{r}'].number_format = '₹#,##0.00'
                if r >= 17:  # Components section
                    ws[f'D{r}'].number_format = '₹#,##0.00'
                    ws[f'E{r}'].number_format = '₹#,##0.00'
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Return response
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="estimate_{estimate.project.name}.xlsx"'
            
            return response
            
        except ImportError:
            return Response(
                {'error': 'Excel generation not available. Install openpyxl: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {'error': f'Excel generation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ComponentViewSet(viewsets.ModelViewSet):
    """API endpoint for components"""
    queryset = Component.objects.all()
    serializer_class = ComponentSerializer
    
    def get_queryset(self):
        queryset = Component.objects.all()
        estimate_id = self.request.query_params.get('estimate_id', None)
        
        if estimate_id:
            queryset = queryset.filter(estimate_id=estimate_id)
        
        return queryset
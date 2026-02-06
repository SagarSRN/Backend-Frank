"""
Enhanced Views - Add these endpoints to apps/estimates/views.py
"""

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.http import FileResponse, HttpResponse
import os

from apps.projects.models import Project
from apps.estimates.models import Estimate, EstimateLineItem, RoomEstimate
from apps.estimates.services_enhanced import (
    generate_detailed_estimate,
    get_detailed_estimate_for_api,
    get_estimate_summary_by_category
)
from apps.estimates.pdf_generator import generate_estimate_pdf


@api_view(['GET'])
def estimate_detailed(request, project_id):
    """
    Get detailed estimate with all line items
    
    GET /api/projects/{id}/estimate-detailed/
    """
    try:
        data = get_detailed_estimate_for_api(project_id)
        
        if not data:
            return Response(
                {'error': 'No estimate found for this project'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        return Response(data)
    
    except Project.DoesNotExist:
        return Response(
            {'error': 'Project not found'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['POST'])
def regenerate_detailed_estimate(request, project_id):
    """
    Regenerate detailed estimate with line items
    
    POST /api/projects/{id}/regenerate-estimate/
    """
    try:
        project = get_object_or_404(Project, id=project_id)
        
        estimate = generate_detailed_estimate(project_id)
        
        if not estimate:
            return Response(
                {'error': 'Could not generate estimate. No rooms found.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        data = get_detailed_estimate_for_api(project_id)
        
        return Response({
            'message': 'Estimate regenerated successfully',
            'data': data
        }, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def estimate_by_category(request, project_id):
    """
    Get cost breakdown by category (Civil, Interior, etc.)
    
    GET /api/projects/{id}/estimate-by-category/
    """
    try:
        category_summary = get_estimate_summary_by_category(project_id)
        
        if not category_summary:
            return Response(
                {'error': 'No estimate found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        total = sum(category_summary.values())
        gst = round(total * 0.18, 2)
        
        return Response({
            'categories': category_summary,
            'subtotal': round(total, 2),
            'gst': gst,
            'grand_total': round(total + gst, 2)
        })
    
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def download_estimate_excel(request, project_id):
    """
    Download estimate as Excel file
    
    GET /api/projects/{id}/estimate/download-excel/
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        project = get_object_or_404(Project, id=project_id)
        data = get_detailed_estimate_for_api(project_id)
        
        if not data:
            return Response(
                {'error': 'No estimate found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estimate"
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"PROJECT ESTIMATE - {project.name}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Project Info
        ws['A3'] = 'Project Type:'
        ws['B3'] = project.project_type
        ws['A4'] = 'Scope:'
        ws['B4'] = project.scope
        ws['A5'] = 'Location:'
        ws['B5'] = project.location
        
        # Headers
        headers = ['Category', 'Item', 'Room', 'Quantity', 'Unit', 'Rate (₹)', 'Amount (₹)']
        header_row = 7
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = header_row + 1
        for item in data['line_items']:
            ws.cell(row=row, column=1, value=item['category'])
            ws.cell(row=row, column=2, value=item['item_name'])
            ws.cell(row=row, column=3, value=item['room'] or '-')
            ws.cell(row=row, column=4, value=item['quantity'])
            ws.cell(row=row, column=5, value=item['unit'])
            ws.cell(row=row, column=6, value=item['rate'])
            ws.cell(row=row, column=7, value=item['amount'])
            row += 1
        
        # Totals
        row += 1
        ws.cell(row=row, column=6, value='Subtotal:').font = Font(bold=True)
        ws.cell(row=row, column=7, value=data['summary']['total_cost']).font = Font(bold=True)
        
        row += 1
        ws.cell(row=row, column=6, value='GST (18%):').font = Font(bold=True)
        ws.cell(row=row, column=7, value=data['summary']['gst']).font = Font(bold=True)
        
        row += 1
        ws.cell(row=row, column=6, value='Grand Total:').font = Font(bold=True, size=14)
        ws.cell(row=row, column=7, value=data['summary']['grand_total']).font = Font(bold=True, size=14)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=estimate_{project.name.replace(" ", "_")}.xlsx'
        
        wb.save(response)
        return response
    
    except Exception as e:
        return Response(
            {'error': f'Error generating Excel: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def dxf_analysis_info(request, project_id):
    """
    Get information about what was detected from DXF file
    
    GET /api/projects/{id}/dxf-analysis/
    """
    from apps.uploads.models import PlanUpload
    from apps.uploads.dxf_processor import get_dxf_info
    from apps.rooms.models import Room
    
    try:
        upload = PlanUpload.objects.filter(project_id=project_id).first()
        
        if not upload:
            return Response(
                {'error': 'No DXF file uploaded for this project'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get DXF file information
        dxf_info = get_dxf_info(upload.file.path)
        
        # Get detected rooms
        rooms = Room.objects.filter(project_id=project_id)
        
        entity_counts = dxf_info.get('entities', {})
        
        return Response({
            'file_info': {
                'file_name': os.path.basename(upload.file.name),
                'file_type': upload.file_type,
                'scale': upload.scale,
                'uploaded_at': upload.uploaded_at,
                'processed': upload.processed,
            },
            'dxf_info': {
                'version': dxf_info.get('version', 'Unknown'),
                'total_entities': dxf_info.get('total_entities', 0),
                'polylines': entity_counts.get('LWPOLYLINE', 0) + entity_counts.get('POLYLINE', 0),
                'text_labels': entity_counts.get('TEXT', 0) + entity_counts.get('MTEXT', 0),
                'lines': entity_counts.get('LINE', 0),
                'circles': entity_counts.get('CIRCLE', 0),
            },
            'detection_results': {
                'total_rooms_detected': rooms.count(),
                'rooms': [
                    {
                        'name': room.name,
                        'area': room.area,
                        'position': {'x': room.x_center, 'y': room.y_center}
                    }
                    for room in rooms
                ]
            }
        })
    
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# Add these URL patterns to apps/estimates/urls.py:
"""
from apps.estimates.views_enhanced import (
    estimate_detailed,
    regenerate_detailed_estimate,
    estimate_by_category,
    download_estimate_excel,
    dxf_analysis_info
)

urlpatterns = [
    # ... existing patterns ...
    
    # Detailed estimate endpoints
    path('projects/<int:project_id>/estimate-detailed/', estimate_detailed),
    path('projects/<int:project_id>/regenerate-estimate/', regenerate_detailed_estimate),
    path('projects/<int:project_id>/estimate-by-category/', estimate_by_category),
    path('projects/<int:project_id>/estimate/download-excel/', download_estimate_excel),
    path('projects/<int:project_id>/dxf-analysis/', dxf_analysis_info),
]
"""
